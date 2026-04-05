import csv
import io
import json
import re
import uuid
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from src.models import Lead, LeadStatus
from src.services.lead_service import lead_service
from src.services.lead_audit_service import lead_audit_service


FIELD_ALIASES: dict[str, list[str]] = {
    "full_name": ["фио", "имя", "клиент", "fullname", "contactname", "name"],
    "phone": ["телефон", "контактныйтелефон", "мобильный", "номер", "phone", "phonenumber", "contactphone"],
    "username": ["username", "telegramusername", "telegram", "tg", "ник", "никнейм"],
    "email": ["email", "e-mail", "почта", "электроннаяпочта", "mail"],
    "address": ["адрес", "жк", "жилойкомплекс", "локация", "address", "complex"],
    "property_type": ["типобъекта", "типобъектанедвижимости", "объект", "недвижимость", "propertytype", "objecttype"],
    "area_sqm": ["площадь", "квм", "квм2", "area", "areasqm"],
    "renovation_type": ["типремонта", "ремонт", "renovation", "renovationtype"],
    "budget": ["бюджет", "стоимость", "цена", "budget", "amount"],
    "deadline": ["срок", "сроки", "дедлайн", "deadline", "duedate"],
    "status": ["статус", "воронка", "stage", "leadstatus"],
}

STANDARD_EXTRACTED_FIELDS = {
    "email": "email",
    "address": "address",
    "property_type": "property_type",
    "area_sqm": "area_sqm",
    "renovation_type": "renovation_type",
    "budget": "budget",
    "deadline": "deadline",
}


class LeadImportService:
    @staticmethod
    async def import_leads_from_file(
        db: AsyncSession,
        org_id: uuid.UUID,
        filename: str,
        file_bytes: bytes,
        source: str = "IMPORT",
        actor_user_id: Optional[uuid.UUID] = None,
    ) -> dict[str, Any]:
        if not file_bytes:
            raise ValueError("Файл пустой.")

        headers, raw_rows = LeadImportService._parse_file(filename=filename, file_bytes=file_bytes)
        mapping = LeadImportService._detect_column_mapping(headers)
        mapping = LeadImportService._enhance_mapping_with_value_heuristics(
            headers=headers,
            rows=raw_rows,
            mapping=mapping,
        )
        detected_columns = {
            field: headers[index]
            for field, index in mapping.items()
            if 0 <= index < len(headers)
        }

        total_rows = 0
        imported = 0
        updated = 0
        skipped = 0
        errors: list[dict[str, Any]] = []

        for row_number, row_values in raw_rows:
            total_rows += 1
            row_dict = LeadImportService._row_to_dict(headers, row_values)

            full_name = LeadImportService._clean_value(
                LeadImportService._get_mapped_value(mapping, "full_name", row_values)
            )
            phone_raw = LeadImportService._clean_value(
                LeadImportService._get_mapped_value(mapping, "phone", row_values)
            )
            username_raw = LeadImportService._clean_value(
                LeadImportService._get_mapped_value(mapping, "username", row_values)
            )
            status_raw = LeadImportService._clean_value(
                LeadImportService._get_mapped_value(mapping, "status", row_values)
            )

            phone = LeadImportService._normalize_phone(phone_raw)
            username = LeadImportService._normalize_username(username_raw)
            lead_status = LeadImportService._parse_status(status_raw) or LeadStatus.NEW

            if not any([full_name, phone, username]):
                skipped += 1
                errors.append({"row": row_number, "reason": "Нет ключевых данных (ФИО / телефон / username)."})
                continue

            extracted_data = LeadImportService._build_extracted_data(
                headers=headers,
                row_values=row_values,
                row_dict=row_dict,
                mapping=mapping,
                full_name=full_name,
                phone=phone,
            )

            resolved_contact = await lead_service.resolve_contact_data(
                db=db,
                org_id=org_id,
                full_name=full_name,
                phone=phone,
                username=username,
                source=source,
            )
            if not full_name:
                full_name = resolved_contact.get("full_name") or full_name
            if not username:
                username = resolved_contact.get("username") or username
            telegram_id = resolved_contact.get("telegram_id")
            if resolved_contact.get("messenger_presence"):
                extracted_data.setdefault("messengers", {}).update(resolved_contact["messenger_presence"])
            if resolved_contact.get("whatsapp_wa_id"):
                extracted_data["whatsapp_wa_id"] = resolved_contact["whatsapp_wa_id"]

            try:
                existing_lead = await LeadImportService._find_existing_lead(
                    db=db,
                    org_id=org_id,
                    phone=phone,
                    username=username,
                    telegram_id=telegram_id,
                )

                if existing_lead:
                    before_state = LeadImportService._snapshot_for_history(existing_lead)
                    changed = LeadImportService._apply_updates_to_existing_lead(
                        lead=existing_lead,
                        full_name=full_name,
                        phone=phone,
                        username=username,
                        telegram_id=telegram_id,
                        status=lead_status,
                        source=source,
                        extracted_data=extracted_data,
                        telegram_lookup_status=resolved_contact.get("telegram_lookup_status"),
                        telegram_lookup_checked_at=resolved_contact.get("telegram_lookup_checked_at"),
                        telegram_lookup_error=resolved_contact.get("telegram_lookup_error"),
                    )

                    if changed:
                        after_state = LeadImportService._snapshot_for_history(existing_lead)
                        changes = lead_audit_service.build_field_changes(before_state, after_state)
                        await lead_audit_service.log_change(
                            db=db,
                            lead=existing_lead,
                            action="updated",
                            source="import",
                            user_id=actor_user_id,
                            changes=changes or {"import_sync": {"old": None, "new": f"row:{row_number}"}},
                        )
                        await db.commit()
                        updated += 1
                    else:
                        skipped += 1
                    continue

                new_lead = Lead(
                    org_id=org_id,
                    telegram_id=telegram_id,
                    full_name=full_name,
                    phone=phone,
                    username=username,
                    status=lead_status.value if isinstance(lead_status, LeadStatus) else str(lead_status),
                    source=source,
                    extracted_data=json.dumps(extracted_data, ensure_ascii=False) if extracted_data else None,
                    telegram_lookup_status=resolved_contact.get("telegram_lookup_status") or "not_checked",
                    telegram_lookup_checked_at=resolved_contact.get("telegram_lookup_checked_at"),
                    telegram_lookup_error=resolved_contact.get("telegram_lookup_error"),
                )
                db.add(new_lead)
                await db.flush()
                create_changes = {
                    "full_name": {"old": None, "new": new_lead.full_name},
                    "phone": {"old": None, "new": new_lead.phone},
                    "username": {"old": None, "new": new_lead.username},
                    "status": {"old": None, "new": new_lead.status},
                    "telegram_lookup_status": {"old": None, "new": new_lead.telegram_lookup_status},
                }
                await lead_audit_service.log_change(
                    db=db,
                    lead=new_lead,
                    action="created",
                    source="import",
                    user_id=actor_user_id,
                    changes=create_changes,
                )
                await db.commit()
                imported += 1
            except IntegrityError:
                await db.rollback()
                skipped += 1
                errors.append({"row": row_number, "reason": "Конфликт уникальности (дубликат)."})
            except Exception as exc:
                await db.rollback()
                skipped += 1
                errors.append({"row": row_number, "reason": f"Ошибка обработки: {exc}"})

        return {
            "total_rows": total_rows,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "detected_columns": detected_columns,
            "errors": errors[:50],
        }

    @staticmethod
    def _parse_file(filename: str, file_bytes: bytes) -> tuple[list[str], list[tuple[int, list[Any]]]]:
        extension = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""

        if extension in {"xlsx"}:
            rows = LeadImportService._read_xlsx_rows(file_bytes)
        elif extension in {"csv", "txt"}:
            rows = LeadImportService._read_csv_rows(file_bytes)
        else:
            raise ValueError("Поддерживаются только форматы .xlsx и .csv")

        headers, data_rows = LeadImportService._extract_header_and_rows(rows)
        return headers, data_rows

    @staticmethod
    def _read_xlsx_rows(file_bytes: bytes) -> list[tuple[int, list[Any]]]:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active
        rows: list[tuple[int, list[Any]]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            rows.append((index, list(row)))
        return rows

    @staticmethod
    def _read_csv_rows(file_bytes: bytes) -> list[tuple[int, list[Any]]]:
        text = None
        for encoding in ("utf-8-sig", "cp1251", "utf-16", "latin-1"):
            try:
                text = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Не удалось прочитать CSV (кодировка не поддерживается).")

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"

        reader = csv.reader(io.StringIO(text), dialect)
        return [(idx, row) for idx, row in enumerate(reader, start=1)]

    @staticmethod
    def _extract_header_and_rows(
        rows: list[tuple[int, list[Any]]]
    ) -> tuple[list[str], list[tuple[int, list[Any]]]]:
        for idx, (_, row_values) in enumerate(rows):
            cleaned = [LeadImportService._clean_value(value) for value in row_values]
            non_empty = sum(1 for value in cleaned if value)
            text_like = sum(1 for value in cleaned if value and re.search(r"[A-Za-zА-Яа-я]", value))

            if non_empty >= 2 and text_like >= 1:
                headers = [value or f"column_{col_idx + 1}" for col_idx, value in enumerate(cleaned)]
                return headers, rows[idx + 1 :]

        raise ValueError("Не удалось определить строку заголовков.")

    @staticmethod
    def _detect_column_mapping(headers: list[str]) -> dict[str, int]:
        candidates: list[tuple[int, str, int]] = []
        for index, header in enumerate(headers):
            normalized = LeadImportService._normalize_text(header)
            if not normalized:
                continue

            for field, aliases in FIELD_ALIASES.items():
                score = LeadImportService._score_header_match(normalized, aliases)
                if score > 0:
                    candidates.append((score, field, index))

        candidates.sort(key=lambda item: item[0], reverse=True)

        mapping: dict[str, int] = {}
        used_indexes: set[int] = set()
        for _, field, index in candidates:
            if field in mapping or index in used_indexes:
                continue
            mapping[field] = index
            used_indexes.add(index)
        return mapping

    @staticmethod
    def _enhance_mapping_with_value_heuristics(
        headers: list[str],
        rows: list[tuple[int, list[Any]]],
        mapping: dict[str, int],
    ) -> dict[str, int]:
        if not rows:
            return mapping

        sampled_rows = rows[:40]
        used_indexes = set(mapping.values())

        def column_values(index: int) -> list[str]:
            values: list[str] = []
            for _, row in sampled_rows:
                if index >= len(row):
                    continue
                cleaned = LeadImportService._clean_value(row[index])
                if cleaned:
                    values.append(cleaned)
            return values

        if "phone" not in mapping:
            best_idx = None
            best_score = 0.0
            for idx in range(len(headers)):
                if idx in used_indexes:
                    continue
                values = column_values(idx)
                if not values:
                    continue
                phone_like = sum(1 for value in values if len(re.sub(r"\D", "", value)) >= 10)
                score = phone_like / len(values)
                if score > best_score:
                    best_score = score
                    best_idx = idx
            if best_idx is not None and best_score >= 0.55:
                mapping["phone"] = best_idx
                used_indexes.add(best_idx)

        if "username" not in mapping:
            best_idx = None
            best_score = 0.0
            for idx in range(len(headers)):
                if idx in used_indexes:
                    continue
                values = column_values(idx)
                if not values:
                    continue
                username_like = 0
                for value in values:
                    test = LeadImportService._normalize_username(value) or ""
                    if re.fullmatch(r"[A-Za-z0-9_]{3,64}", test):
                        username_like += 1
                score = username_like / len(values)
                if score > best_score:
                    best_score = score
                    best_idx = idx
            if best_idx is not None and best_score >= 0.6:
                mapping["username"] = best_idx
                used_indexes.add(best_idx)

        if "full_name" not in mapping:
            best_idx = None
            best_score = 0.0
            for idx in range(len(headers)):
                if idx in used_indexes:
                    continue
                values = column_values(idx)
                if not values:
                    continue
                name_like = 0
                for value in values:
                    words = [word for word in re.split(r"\s+", value) if word]
                    has_letters = bool(re.search(r"[A-Za-zА-Яа-я]", value))
                    if has_letters and len(words) >= 2:
                        name_like += 1
                score = name_like / len(values)
                if score > best_score:
                    best_score = score
                    best_idx = idx
            if best_idx is not None and best_score >= 0.5:
                mapping["full_name"] = best_idx
                used_indexes.add(best_idx)

        return mapping

    @staticmethod
    def _score_header_match(normalized_header: str, aliases: list[str]) -> int:
        best = 0
        for alias in aliases:
            normalized_alias = LeadImportService._normalize_text(alias)
            if not normalized_alias:
                continue

            if normalized_header == normalized_alias:
                best = max(best, 100 + len(normalized_alias))
                continue

            if len(normalized_alias) >= 2 and normalized_alias in normalized_header:
                best = max(best, 75 + len(normalized_alias))
                continue

            if len(normalized_header) >= 5 and normalized_header in normalized_alias:
                best = max(best, 50 + len(normalized_header))
        return best

    @staticmethod
    def _row_to_dict(headers: list[str], row_values: list[Any]) -> dict[str, str]:
        row_dict: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = LeadImportService._clean_value(row_values[index] if index < len(row_values) else None)
            row_dict[header] = value or ""
        return row_dict

    @staticmethod
    def _build_extracted_data(
        headers: list[str],
        row_values: list[Any],
        row_dict: dict[str, str],
        mapping: dict[str, int],
        full_name: Optional[str],
        phone: Optional[str],
    ) -> dict[str, Any]:
        extracted_data: dict[str, Any] = {}

        for field_name, extracted_key in STANDARD_EXTRACTED_FIELDS.items():
            value = LeadImportService._clean_value(
                LeadImportService._get_mapped_value(mapping, field_name, row_values)
            )
            if not value:
                continue
            if field_name == "area_sqm":
                parsed_number = LeadImportService._parse_number(value)
                extracted_data[extracted_key] = parsed_number if parsed_number is not None else value
            else:
                extracted_data[extracted_key] = value

        if full_name:
            extracted_data.setdefault("client_name", full_name)
        if phone:
            extracted_data.setdefault("phone", phone)

        mapped_indexes = set(mapping.values())
        for index, header in enumerate(headers):
            if index in mapped_indexes:
                continue
            raw_value = row_dict.get(header, "").strip()
            if raw_value:
                extracted_data[header] = raw_value

        return extracted_data

    @staticmethod
    async def _find_existing_lead(
        db: AsyncSession,
        org_id: uuid.UUID,
        phone: Optional[str],
        username: Optional[str],
        telegram_id: Optional[int],
    ) -> Optional[Lead]:
        conditions = []
        if phone:
            conditions.append(Lead.phone == phone)
        if username:
            conditions.append(func.lower(Lead.username) == username.lower())
        if telegram_id:
            conditions.append(Lead.telegram_id == telegram_id)

        if not conditions:
            return None

        result = await db.execute(
            select(Lead)
            .where(Lead.org_id == org_id)
            .where(or_(*conditions))
            .limit(1)
        )
        return result.scalar_one_or_none()

    @staticmethod
    def _apply_updates_to_existing_lead(
        lead: Lead,
        full_name: Optional[str],
        phone: Optional[str],
        username: Optional[str],
        telegram_id: Optional[int],
        status: LeadStatus,
        source: str,
        extracted_data: dict[str, Any],
        telegram_lookup_status: Optional[str],
        telegram_lookup_checked_at: Any,
        telegram_lookup_error: Optional[str],
    ) -> bool:
        changed = False

        if full_name and not lead.full_name:
            lead.full_name = full_name
            changed = True

        if phone and not lead.phone:
            lead.phone = phone
            changed = True

        if username and not lead.username:
            lead.username = username
            changed = True

        if telegram_id and not lead.telegram_id:
            lead.telegram_id = telegram_id
            changed = True

        if source and not lead.source:
            lead.source = source
            changed = True

        if telegram_lookup_status and lead.telegram_lookup_status != telegram_lookup_status:
            lead.telegram_lookup_status = telegram_lookup_status
            changed = True

        if telegram_lookup_checked_at and lead.telegram_lookup_checked_at != telegram_lookup_checked_at:
            lead.telegram_lookup_checked_at = telegram_lookup_checked_at
            changed = True

        if lead.telegram_lookup_error != telegram_lookup_error:
            lead.telegram_lookup_error = telegram_lookup_error
            changed = True

        if status and (not lead.status or str(lead.status).upper() == LeadStatus.NEW.value):
            if str(lead.status).upper() != status.value:
                lead.status = status.value
                changed = True

        existing_extracted: dict[str, Any] = {}
        if lead.extracted_data:
            try:
                existing_extracted = json.loads(lead.extracted_data)
            except Exception:
                existing_extracted = {}

        merged = dict(existing_extracted)
        merged.update(extracted_data)

        if merged != existing_extracted:
            lead.extracted_data = json.dumps(merged, ensure_ascii=False)
            changed = True

        return changed

    @staticmethod
    def _snapshot_for_history(lead: Lead) -> dict[str, Any]:
        return {
            "full_name": lead.full_name,
            "phone": lead.phone,
            "username": lead.username,
            "telegram_id": lead.telegram_id,
            "status": lead.status,
            "source": lead.source,
            "extracted_data": lead.extracted_data,
            "telegram_lookup_status": lead.telegram_lookup_status,
            "telegram_lookup_error": lead.telegram_lookup_error,
        }

    @staticmethod
    def _get_mapped_value(mapping: dict[str, int], field_name: str, row_values: list[Any]) -> Any:
        index = mapping.get(field_name)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    @staticmethod
    def _clean_value(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        cleaned = str(value).strip()
        if not cleaned:
            return None
        return cleaned

    @staticmethod
    def _normalize_text(value: str) -> str:
        text = (value or "").lower().replace("ё", "е")
        text = re.sub(r"[^a-zа-я0-9]+", "", text, flags=re.IGNORECASE)
        return text

    @staticmethod
    def _normalize_phone(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        digits = re.sub(r"\D", "", value)
        if not digits:
            return None
        if len(digits) == 11 and digits.startswith("8"):
            digits = "7" + digits[1:]
        if len(digits) == 10:
            digits = "7" + digits
        if len(digits) < 10:
            return value.strip()
        return f"+{digits}"

    @staticmethod
    def _normalize_username(value: Optional[str]) -> Optional[str]:
        if not value:
            return None

        username = value.strip()
        username = re.sub(r"^https?://t\.me/", "", username, flags=re.IGNORECASE)
        username = username.strip().lstrip("@")
        username = re.sub(r"\s+", "", username)
        return username or None

    @staticmethod
    def _parse_number(value: str) -> Optional[float]:
        normalized = value.replace(" ", "").replace(",", ".")
        match = re.search(r"-?\d+(?:\.\d+)?", normalized)
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None

    @staticmethod
    def _parse_status(value: Optional[str]) -> Optional[LeadStatus]:
        if not value:
            return None

        normalized = LeadImportService._normalize_text(value)
        if not normalized:
            return None

        direct_by_value = {status.value.lower(): status for status in LeadStatus}
        if normalized in direct_by_value:
            return direct_by_value[normalized]

        status_rules: list[tuple[list[str], LeadStatus]] = [
            (["new", "нов"], LeadStatus.NEW),
            (["consult", "консультац"], LeadStatus.CONSULTING),
            (["follow", "догрев", "фоллоу"], LeadStatus.FOLLOW_UP),
            (["qualif", "квалиф"], LeadStatus.QUALIFIED),
            (["measur", "замер"], LeadStatus.MEASUREMENT),
            (["estimate", "смет"], LeadStatus.ESTIMATE),
            (["contract", "контракт"], LeadStatus.CONTRACT),
            (["won", "успех", "выиг"], LeadStatus.WON),
            (["lost", "проиг", "отказ"], LeadStatus.LOST),
            (["spam", "спам"], LeadStatus.SPAM),
        ]
        for keywords, status in status_rules:
            if any(keyword in normalized for keyword in keywords):
                return status

        return None


lead_import_service = LeadImportService()
