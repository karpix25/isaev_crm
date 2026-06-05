from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.services.estimates.types import Estimate, EstimateLine


HEADER_ROW = 7


def export_isaev_estimate_xlsx(estimate: Estimate, output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(estimate.address)

    styles = _styles()
    _write_header(ws, estimate)
    row = HEADER_ROW + 2

    rough_total_row = None
    clean_total_row = None
    for section_index, section in enumerate(estimate.sections):
        ws.cell(row=row, column=1, value=section.title)
        _fill_row(ws, row, 1, 6, styles["major"])
        row += 1

        for subsection in section.subsections:
            start = row + 1
            _write_subsection_row(ws, row, subsection.section_no, subsection.title, subsection.total, styles)
            row += 1
            for item in subsection.lines:
                _write_line(ws, row, item, styles)
                row += 1
            _write_formula(ws, row - len(subsection.lines) - 1, start, row - 1)

        total_label = "ИТОГО по черновым работам" if section_index == 0 else "ИТОГО по чистовым работам"
        total_row = row
        ws.cell(row=row, column=1, value=total_label)
        ws.cell(row=row, column=6, value=_num(section.total))
        _fill_row(ws, row, 1, 6, styles["total"])
        row += 1

        ws.cell(row=row, column=1, value="Цена за м2")
        row += 1
        ws.cell(row=row, column=1, value=f"Скидка {_discount_percent(estimate)}%")
        row += 1

        discounted_label = (
            "ИТОГО с учётом скидки* по черновым работам"
            if section_index == 0
            else "ИТОГО с учётом скидки* по чистовым работам"
        )
        ws.cell(row=row, column=1, value=discounted_label)
        ws.cell(row=row, column=6, value=_num(section.total * (Decimal("1") - estimate.discount_rate)))
        _fill_row(ws, row, 1, 6, styles["total"])
        if section_index == 0:
            rough_total_row = row
        else:
            clean_total_row = row
        row += 2

    ws.cell(row=row, column=1, value="ВСЕГО с учётом скидки* по черновым и чистовым работам")
    ws.cell(row=row, column=6, value=_num(estimate.discounted_total))
    _fill_row(ws, row, 1, 6, styles["grand"])
    row += 2

    for note in estimate.notes:
        ws.cell(row=row, column=1, value=note)
        row += 1

    _format_sheet(ws)
    wb.save(path)
    return path


def _write_header(ws, estimate: Estimate) -> None:
    ws.cell(row=1, column=1, value="СМЕТА")
    ws.cell(row=2, column=1, value=f"по адресу: {estimate.address}")
    ws.cell(row=3, column=1, value=estimate.contract_term_text)
    ws.cell(row=4, column=1, value=estimate.contract_text)
    if estimate.valid_until:
        ws.cell(row=5, column=1, value=f"Стоимость действительна до {estimate.valid_until}")
    headers = ["№", "Наименование работ", "Цена", "Кол-во", "Ед.изм.", "Стоимость"]
    for col, title in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=title)
        ws.cell(row=HEADER_ROW, column=col).font = Font(bold=True)


def _write_subsection_row(ws, row: int, section_no: int, title: str, total: Decimal, styles: dict) -> None:
    ws.cell(row=row, column=1, value=section_no)
    ws.cell(row=row, column=2, value=title)
    ws.cell(row=row, column=4, value="Итого")
    ws.cell(row=row, column=6, value=_num(total))
    _fill_row(ws, row, 1, 6, styles["subsection"])


def _write_line(ws, row: int, item: EstimateLine, styles: dict) -> None:
    values = [item.line_no, item.name, _num(item.price), _num(item.quantity), item.unit, _num(item.amount)]
    for col, value in enumerate(values, 1):
        ws.cell(row=row, column=col, value=value)
    _fill_row(ws, row, 1, 6, styles["body"])


def _write_formula(ws, row: int, start: int, end: int) -> None:
    if end >= start:
        ws.cell(row=row, column=6, value=f"=SUM(F{start}:F{end})")


def _format_sheet(ws) -> None:
    widths = [8, 70, 13, 12, 12, 16, 4]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in (3, 4, 6):
        for cell in ws[get_column_letter(col)]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    ws.freeze_panes = "A8"


def _fill_row(ws, row: int, start_col: int, end_col: int, style: dict) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = style["fill"]
        cell.border = style["border"]
        if style.get("font"):
            cell.font = style["font"]


def _styles() -> dict:
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "major": {"fill": PatternFill("solid", fgColor="E7E6E6"), "border": border, "font": Font(bold=True)},
        "subsection": {"fill": PatternFill("solid", fgColor="D9EAD3"), "border": border, "font": Font(bold=True)},
        "total": {"fill": PatternFill("solid", fgColor="FCE4D6"), "border": border, "font": Font(bold=True)},
        "grand": {"fill": PatternFill("solid", fgColor="D9EAF7"), "border": border, "font": Font(bold=True)},
        "body": {"fill": PatternFill("solid", fgColor="FFFFFF"), "border": border},
    }


def _num(value: Decimal) -> float:
    return float(value)


def _discount_percent(estimate: Estimate) -> str:
    percent = estimate.discount_rate * Decimal("100")
    return str(percent.quantize(Decimal("1"))) if percent == percent.to_integral() else str(percent)


def _sheet_title(address: str) -> str:
    title = "Смета " + address.replace("/", " ").replace("\\", " ")
    return title[:31]
